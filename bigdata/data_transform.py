# encoding=utf-8
import sys

reload(sys)
sys.setdefaultencoding('utf-8')
from openpyxl import load_workbook
from bigdata.mongo_util import insert_one

if __name__ == '__main__':
    wb = load_workbook(filename='/Users/alex/Downloads/data.xlsx')
    worksheet1 = wb.get_sheet_by_name("电话")
    for row_index in range(2, worksheet1.max_row):
        json = {
            "id": 'tel_' + worksheet1['A%s' % row_index].value,
            "sent_at": worksheet1['B%s' % row_index].value,
            "sent_from": worksheet1['E%s' % row_index].value,
            "gender": worksheet1['F%s' % row_index].value,
            "district": worksheet1['D%s' % row_index].value,
            "address": worksheet1['G%s' % row_index].value,
            "title": worksheet1['U%s' % row_index].value,
            "content": worksheet1['J%s' % row_index].value,
            "help_category": worksheet1['N%s' % row_index].value,
            "content_category": worksheet1['P%s' % row_index].value,
            "handle_at": worksheet1['R%s' % row_index].value
        }
        insert_one(json)
    print "finsihed worksheet 1"

    worksheet2 = wb.get_sheet_by_name("信件")
    for row_index in range(2, worksheet2.max_row):
        json = {
            "id": 'mail_' + worksheet2['A%s' % row_index].value,
            "sent_at": worksheet2['M%s' % row_index].value,
            "sent_from": worksheet2['D%s' % row_index].value,
            "gender": worksheet2['E%s' % row_index].value,
            "district": worksheet2['C%s' % row_index].value,
            "address": worksheet2['F%s' % row_index].value,
            "title": worksheet2['H%s' % row_index].value,
            "content": worksheet2['L%s' % row_index].value,
            "help_category": worksheet2['R%s' % row_index].value,
            "content_category": worksheet2['T%s' % row_index].value,
            "handle_at": worksheet2['W%s' % row_index].value
        }
        insert_one(json)
    print "finsihed worksheet 2"

    worksheet3 = wb.get_sheet_by_name("短信")
    for row_index in range(2, worksheet3.max_row):
        json = {
            "id": 'sms_' + worksheet3['A%s' % row_index].value,
            "sent_at": worksheet3['K%s' % row_index].value,
            "district": worksheet3['C%s' % row_index].value,
            "address": worksheet3['F%s' % row_index].value,
            "title": worksheet3['H%s' % row_index].value,
            "content": worksheet3['J%s' % row_index].value,
            "help_category": worksheet3['O%s' % row_index].value,
            "content_category": worksheet3['Q%s' % row_index].value,
            "handle_at": worksheet1['T%s' % row_index].value
        }
        insert_one(json)
    print "finsihed worksheet 3"
